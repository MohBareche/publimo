import tkinter as tk
import docx2pdf
import os
import re
import glob
import shutil
import win32com.client

from tkinter import ttk, messagebox as mb, filedialog as fd, IntVar
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from pathlib import Path
from tkcalendar import Calendar, DateEntry
from PyPDF2 import PdfMerger, PdfReader

global data_name
global ent_list
global soum_list
global adj_list

to_list = []
from_list = []


def select_data_file():
    filetypes = (
        ('Fichier Excel', '*.xlsx'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Choisir une base de données',
        initialdir='./data',
        filetypes=filetypes
    )
    data_name = Path(filename)
    global wb
    wb = load_workbook(data_name)

    # liste chargés de projet
    ws_charg_proj = wb['Chargés de projet']
    list_charg_proj = []
    for cell in ws_charg_proj['B'][1:]:
        if cell.value != 'None':
            list_charg_proj.append(cell.value)
        lbl_message.grid(row=0, column=2)
        cmb_nom_charg_projet['values'] = list_charg_proj
        cmb_nom_charg_projet.configure(state='readonly')

    # liste gestionnaires
    ws_gest = wb['Gestionnaires']
    list_gestionnaires = []
    for cell in ws_gest['A'][1:]:
        if cell.value != 'None':
            list_gestionnaires.append(cell.value)
        cmb_nom_gestionnaire['values'] = list_gestionnaires
        cmb_nom_gestionnaire.current(0)
        cmb_nom_gestionnaire.configure(state='readonly')

    # liste secrétaires
    list_secret = []
    for cell in ws_gest['E'][1:]:
        if cell.value != 'None':
            list_secret.append(cell.value)
        cmb_secretaire['values'] = list_secret
        cmb_secretaire.current(0)
        cmb_secretaire.configure(state='readonly')

    lbl_message.configure(
        text='Base de données chargée avec succès...', width=50, relief='groove', fg='lime')


def load_data():
    excel_filename = 'data/Registre des entrepreneurs.xlsx'
    path_current = os.getcwd()
    path = f"{path_current}\{excel_filename}"

    wb = load_workbook(path)
    sheet = wb.active

    list_values = list(sheet.values)
    cols = list_values[0]

    tree = ttk.Treeview(root, columns=cols, show='headings')
    tree.pack(expand=True, fill='y')

    for col_name in cols:
        tree.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        tree.insert('', tk.END, values=value_tuple)

    tree.column('Ville', width=160, anchor=tk.CENTER)
    tree.column('Représentant', width=130)
    tree.column('Code Postal', width=100, anchor=tk.CENTER)
    tree.column('Civilité', width=80, anchor=tk.CENTER)
    tree.column('Fonction', width=150)


def moveTo(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())


def move_adj(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='normal')


def back_adj(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())
        btn_adj_1.configure(state='normal')
        btn_adj_2.configure(state='disabled')


def soum_to_adj(e):
    if not adj_list.get(0, tk.END):
        btn_adj_1.configure(state='normal')
    else:
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='normal')


def adj_to_soum(e):
    if not adj_list.get(0, tk.END):
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='disabled')
    else:
        btn_adj_2.configure(state='normal')


def move_all(f_list, t_list):
    all_items = f_list.get(0, tk.END)
    f_list.delete(0, tk.END)
    for item in all_items:
        t_list.insert(tk.END, item)


def dbl_moveTo(e):
    ind_list = ent_list.curselection()
    if ind_list:
        ind = ind_list[0]
        val = ent_list.get(ind)
        ent_list.delete(ind)
        soum_list.insert(tk.END, val)


def dbl_moveBack(e):
    ind_list = soum_list.curselection()
    if ind_list:
        ind = ind_list[0]
        val = soum_list.get(ind)
        soum_list.delete(ind)
        ent_list.insert(tk.END, val)


def show_list_ent(e):
    ent_list.delete(0, tk.END)
    soum_list.delete(0, tk.END)
    nom_charg_proj = cmb_nom_charg_projet.get()
    data = wb['Chargés de projet']

    for row in data.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value == nom_charg_proj:
                global discipline
                specialite = data.cell(row=cell.row, column=4).value
                discipline = specialite

                if specialite == 'Voirie':
                    sheet_voirie = wb['Voirie']
                    m_row = sheet_voirie.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = sheet_voirie.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)

                if specialite == 'Bâtiment':
                    list_ent_bat = wb['Bâtiment']
                    m_row = list_ent_bat.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = list_ent_bat.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)

                if specialite == 'APA':
                    list_ent_apa = wb['Paysage']
                    m_row = list_ent_apa.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = list_ent_apa.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)


def select_remerc_file():
    global doc_remerc_name
    filetypes = (
        ('Fichier Word', '*.docx'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Choisir un gabarit',
        initialdir='./gabarits',
        filetypes=filetypes
    )
    doc_remerc_name = Path(filename).name
    if doc_remerc_name:
        lbl_remerc.configure(fg='green', text='Gabarit remerciement (OK)')
        return doc_remerc_name


def select_octroi_file():
    global doc_octroi_name
    filetypes = (
        ('Fichier Word', '*.docx'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Choisir un gabarit',
        initialdir='./gabarits',
        filetypes=filetypes
    )
    doc_octroi_name = Path(filename).name
    if doc_octroi_name:
        lbl_octroi.configure(fg='green', text="Gabarit d'octroi (OK)")
        return doc_octroi_name


def select_pv_ouverture_file():
    global doc_pv_ouvert_name
    filetypes = (
        ('Fichier PDF', '*.pdf'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title="Sélectionner le PV d'ouverture",
        initialdir='./pv',
        filetypes=filetypes
    )
    doc_pv_ouvert_name = Path(filename).name
    if doc_pv_ouvert_name:
        lbl_pv_ouvert.configure(
            fg='green', text="Procès verbal d'ouverture (OK)")
        return doc_pv_ouvert_name


def select_pv_ca_file():
    global doc_pv_ca_name
    filetypes = (
        ('Fichier Word', '*.doc'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Sélectionner le PV du CA',
        initialdir='./pv',
        filetypes=filetypes
    )
    doc_pv_ca_name = Path(filename).name
    if doc_pv_ca_name:
        lbl_pv_ca.configure(fg='green', text="Procès verbal CA (OK)")
        return doc_pv_ca_name


def select_redac():
    global nom_redac
    if var_redac.get() == 0:
        cmb_secretaire.configure(state='readonly')
        nom_redac = cmb_secretaire.get()

    if var_redac.get() == 1:
        cmb_secretaire.configure(state='disabled')
        nom_redac = cmb_nom_charg_projet.get()
    return nom_redac


def get_secret_name(e):
    nom_redac = cmb_secretaire.get()
    return nom_redac


def initiales_gest(nom):
    cap = nom.split(' ')
    init = cap[0][0] + cap[1][0]
    return init


def initiales_redac(nom):
    cap = nom.split(' ')
    init = cap[0][0] + cap[1][0]
    return init.lower()


def gener_remerc():
    path = f'./gabarits/{doc_remerc_name}'
    doc = DocxTemplate(path)
    compagnies = {}
    ws = wb[discipline]
    for row in ws.iter_rows(min_row=2, values_only=True):
        company_name = row[0]
        company_data = {
            "nom_de_compagnie": row[0],
            "adresse": row[1],
            "ville": row[2],
            "code_postal": row[3],
            "courriel": row[4],
            "representant": row[5],
            "civilite": row[6],
            "fonction": row[7]
        }
        compagnies[company_name] = company_data

    ws_gestionnaires = wb['Gestionnaires']
    date = entry_cal.get()
    titre_projet = entry_titre_projet.get()
    num_projet = entry_num_projet.get()
    nom_gest = cmb_nom_gestionnaire.get()
    init_gest = initiales_gest(nom_gest)
    
    if var_redac.get() == 0:
        nom_redac = cmb_secretaire.get()
    if var_redac.get() == 1:
        nom_redac = cmb_nom_charg_projet.get()
    
    init_redac = initiales_redac(nom_redac)

    for row in ws_gestionnaires.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == nom_gest:
                titre_gest = ws_gestionnaires.cell(
                    row=cell.row, column=2).value
                fonction_gest = ws_gestionnaires.cell(
                    row=cell.row, column=3).value

    pathDOC = './output/remerciement/DOC'

    isExist = os.path.exists(pathDOC)
    if not isExist:
        os.makedirs(pathDOC)

    for ent in list(soum_list.get(0, tk.END)):
        doc.render({
            "date": date,
            "titre": titre_projet,
            "num_contrat": num_projet,
            "nom_gestionnaire": nom_gest,
            "titre_gest": titre_gest,
            "fonction_gest": fonction_gest,
            "init_gest": init_gest,
            "init_redac": init_redac,
            "civilite": compagnies[ent]['civilite'],
            "representant": compagnies[ent]['representant'],
            "nom_de_compagnie": compagnies[ent]['nom_de_compagnie'],
            "adresse": compagnies[ent]['adresse'],
            "ville": compagnies[ent]['ville'],
            "code_postal": compagnies[ent]['code_postal'],
            "courriel": compagnies[ent]['courriel']
        })
        nom_comp = f'{compagnies[ent]["nom_de_compagnie"]}'
        nom_fichier = f"Lettre de remerciement - {nom_comp}.docx"

        doc.save(f'{pathDOC}/{nom_fichier}')

    docx2pdf.convert(pathDOC, '.')

    pv = f"./pv/{doc_pv_ouvert_name}"
    pdf_pv = open(pv, 'rb')

    pdfs = glob.glob('*.pdf')

    for pdf in pdfs:
        merger = PdfMerger()
        merger.append(pdf)
        merger.append(pdf_pv)
        name = pdf.split(".")[0]
        merger.write(f"{name}_fin.pdf")
        merger.close()

    for f in glob.glob('./*_fin.pdf'):
        pathPDF = './output/remerciement/PDF'
        isExist = os.path.exists(pathPDF)
        if not isExist:
            os.makedirs(pathPDF)
        shutil.move(f, pathPDF)

    for f in os.listdir('./'):
        if f.endswith('.pdf'):
            os.remove(f)

    mb.showinfo(title='Confirmation',
                message='Publipostage des lettres de remerciement réalisé avec succès.')


def gener_octroi():
    path_gabarit = f'./gabarits/{doc_octroi_name}'
    doc = DocxTemplate(path_gabarit)
    compagnies = {}
    ws = wb[discipline]
    for row in ws.iter_rows(min_row=2, values_only=True):
        company_name = row[0]
        company_data = {
            "nom_de_compagnie": row[0],
            "adresse": row[1],
            "ville": row[2],
            "code_postal": row[3],
            "courriel": row[4],
            "representant": row[5],
            "civilite": row[6],
            "fonction": row[7]
        }
        compagnies[company_name] = company_data

        ws_gestionnaires = wb['Gestionnaires']

    pv_ca = f"./pv/{doc_pv_ca_name}"
    shutil.move(pv_ca, './')
    
    filename = doc_pv_ca_name
    filenamePDF = filename.split('.')[0]
    path = os.getcwd()
    in_file = f"{path}\{filename}"
    out_file = f"{path}\{filenamePDF}"
    
    wdFormatPDF = 17
    word = win32com.client.Dispatch('Word.Application')
    doc_doc = word.Documents.Open(in_file)
    doc_doc.SaveAs(out_file, FileFormat=wdFormatPDF)
    doc_doc.Close()
    word.Quit()
    shutil.move(in_file, './pv')
    
    reader = PdfReader(f"{out_file}.pdf")
    texte =reader.pages[0].extract_text()
    resolution = re.search(r"CA[\d]{2}\s[\d]{2}\s[\d]{2,4}", texte).group()
    date_resolution = re.search(r"[\d]{1,2}\s(?:janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s[\d]{4}", texte).group()
    
    date = entry_cal.get()
    titre_projet = entry_titre_projet.get()
    num_projet = entry_num_projet.get()
    nom_gest = cmb_nom_gestionnaire.get()
    charg_projet = cmb_nom_charg_projet.get()
    init_gest = initiales_gest(nom_gest)
    
    if var_redac.get() == 0:
        nom_redac = cmb_secretaire.get()
    if var_redac.get() == 1:
        nom_redac = cmb_nom_charg_projet.get()
        
    init_redac = initiales_redac(nom_redac)

    ws_charg_proj = wb['Chargés de projet']
    for row in ws_charg_proj.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value == charg_projet:
                civ_charge_proj = ws_charg_proj.cell(
                    row=cell.row, column=1).value
                nom_charge_projet = ws_charg_proj.cell(
                    row=cell.row, column=2).value
                tel_charge_projet = ws_charg_proj.cell(
                    row=cell.row, column=3).value

    ws_gestionnaires = wb['Gestionnaires']
    for row in ws_gestionnaires.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == nom_gest:
                titre_gest = ws_gestionnaires.cell(
                    row=cell.row, column=2).value
                fonction_gest = ws_gestionnaires.cell(
                    row=cell.row, column=3).value

    pathDOC = './output/octroi/DOC'

    isExist = os.path.exists(pathDOC)
    if not isExist:
        os.makedirs(pathDOC)
    for ent in list(adj_list.get(0, tk.END)):
        doc.render({
            "date": date,
            "titre": titre_projet,
            "num_contrat": num_projet,
            "nom_gestionnaire": nom_gest,
            "titre_gest": titre_gest,
            "fonction_gest": fonction_gest,
            "init_gest": init_gest,
            "init_redac": init_redac,
            "resolution": resolution,
            "date_resolution": date_resolution,
            "civ_charge_projet": civ_charge_proj,
            "nom_charge_projet": nom_charge_projet,
            "tel_charge_projet": tel_charge_projet,
            "civilite": compagnies[ent]['civilite'],
            "representant": compagnies[ent]['representant'],
            "nom_de_compagnie": compagnies[ent]['nom_de_compagnie'],
            "adresse": compagnies[ent]['adresse'],
            "ville": compagnies[ent]['ville'],
            "code_postal": compagnies[ent]['code_postal'],
            "courriel": compagnies[ent]['courriel']
        })
        global nom_comp_adj
        nom_comp_adj = f'{compagnies[ent]["nom_de_compagnie"]}'
        nom_fichier_doc = f"Lettre d'adjudication_{nom_comp_adj}.docx"
        doc.save(f'{pathDOC}/{nom_fichier_doc}')

    docx2pdf.convert(pathDOC, '.')

    pdfs = glob.glob('*.pdf')

    pdfs = [f for f in os.listdir() if f.endswith(".pdf")]

    merger = PdfMerger()

    for pdf in pdfs:
        merger.append(open(pdf, 'rb'))

    pathPDF = './output/octroi/PDF'
    isExist = os.path.exists(pathPDF)
    if not isExist:
        os.makedirs(pathPDF)

    nom_fichier_pdf = f"{pathPDF}/Lettre d'adjudication_{num_projet}.pdf"
    with open(nom_fichier_pdf, 'wb') as fout:
        merger.write(fout)
        merger.close()

    for f in os.listdir('./'):
        if f.endswith('.pdf'):
            os.remove(f)

    mb.showinfo(title='Confirmation',
                message="Publipostage de la lettre d'octroi réalisé avec succès.")


root = tk.Tk()
style = ttk.Style()
style.theme_use('clam')
style.configure('Treeview.Heading', background='green2', font=('Bold'))
#  center window in screen
root_width = 1100
root_height = 650
screen_with = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
pos_x = int((screen_with - root_width) // 2)
pos_y = int((screen_height - root_height) // 2)

#  differents config window app
root.title("Publipostage lettres de remerciement et lettre d'octroi")
root.geometry(f"{root_width}x{root_height}+{pos_x}+{pos_y}")
root.resizable(0, 0)

# chargement des données
frame_data = tk.LabelFrame(
    root, text='Choix de la base de données', fg='blue')
frame_data.grid(row=0, column=0, sticky='ew', padx=20, pady=5)

lbl_load_data = tk.Label(frame_data, text='Choisir un fichier')
lbl_load_data.grid(row=0, column=0)
btn_load_data = tk.Button(
    frame_data, text='Sélectionner...', command=select_data_file)
btn_load_data.grid(row=0, column=1)

lbl_message = tk.Label(frame_data, fg='yellow', bg='black',
                       text="Choisir d'abord la base de données pour commencer.", width=50, relief='groove',)
lbl_message.grid(row=0, column=2)

for widget in frame_data.winfo_children():
    widget.grid_configure(padx=10, pady=10)


# informations sur le projet
frame_info_projet = tk.LabelFrame(
    root, text='Informations sur le projet', fg='blue')
frame_info_projet.grid(row=1, column=0, sticky='ew', padx=20, pady=5)

lbl_titre_projet = tk.Label(frame_info_projet, text='Titre du projet')
lbl_titre_projet.grid(row=0, column=0)
entry_titre_projet = tk.Entry(frame_info_projet, width=100)
entry_titre_projet.grid(row=0, column=1)

lbl_num_projet = tk.Label(frame_info_projet, text='Numéro de projet')
lbl_num_projet.grid(row=0, column=2)
entry_num_projet = tk.Entry(frame_info_projet)
entry_num_projet.grid(row=0, column=3)

for widget in frame_info_projet.winfo_children():
    widget.grid_configure(padx=10, pady=10)

frame_info_charg_proj_sign_date = tk.LabelFrame(
    root, text='Informations diverses [ Chargé de projet, signataire, rédacteur et date de rédaction ]', fg='blue')
frame_info_charg_proj_sign_date.grid(
    row=2, column=0, sticky='ew', padx=20, pady=5)

# informations sur le chargé de projet
lbl_nom_charg_projet = tk.Label(
    frame_info_charg_proj_sign_date, text='Chargé(e) de projet')
lbl_nom_charg_projet.grid(row=0, column=0, sticky='s')
cmb_nom_charg_projet = ttk.Combobox(frame_info_charg_proj_sign_date, width=25)
cmb_nom_charg_projet.grid(row=1, column=0, sticky='n')
cmb_nom_charg_projet.bind("<<ComboboxSelected>>", show_list_ent)

# informations sur le signataire (gestionnaire)
lbl_nom_gestionnaire = tk.Label(
    frame_info_charg_proj_sign_date, text='Signataire (Gestionnaire)')
lbl_nom_gestionnaire.grid(row=0, column=1, sticky='s')
cmb_nom_gestionnaire = ttk.Combobox(frame_info_charg_proj_sign_date, width=25)
cmb_nom_gestionnaire.grid(row=1, column=1, sticky='n')

# date de rédaction
lbl_date = tk.Label(frame_info_charg_proj_sign_date, text="Date de rédaction")
lbl_date.grid(row=0, column=2, sticky='s')
entry_cal = DateEntry(frame_info_charg_proj_sign_date,
                      width=16, background="magenta3", foreground="white", bd=2)
entry_cal.grid(row=1, column=2, sticky='n')

# rédacteur
frm_redacteur = tk.LabelFrame(frame_info_charg_proj_sign_date)
frm_redacteur.grid(row=0, column=3, rowspan=2, padx=10, pady=10)
lbl_redacteur = tk.Label(frm_redacteur, text='Rédacteur')
lbl_redacteur.grid(row=0, column=0, rowspan=2)

var_redac = IntVar(None, 0)
rbtn_red = tk.Radiobutton(frm_redacteur, text='Secrétaire',
                          variable=var_redac, value=0, command=select_redac)
rbtn_red.grid(row=0, column=1, sticky='w', padx=10, pady=5)

rbtn_red = tk.Radiobutton(frm_redacteur, text='Chargé(e) de projet',
                          variable=var_redac, value=1, command=select_redac)
rbtn_red.grid(row=1, column=1, sticky='w', padx=10, pady=5)

cmb_secretaire = ttk.Combobox(frm_redacteur, width=25)
cmb_secretaire.grid(row=0, column=2, sticky='w', padx=10)

for widget in frame_info_charg_proj_sign_date.winfo_children():
    widget.grid_configure(padx=15, pady=5)


# ************************************************************************************************
# Frame information soumissionnaires
frame_soumission = tk.LabelFrame(
    root, text='Informations sur les soumissionnaires', fg='blue')
frame_soumission.grid(row=3, column=0, sticky='ew', padx=20, pady=5)

lbl_list_ent = tk.Label(frame_soumission, text='Liste des entrepreneurs')
lbl_list_ent.grid(row=0, column=0)
ent_list = tk.Listbox(frame_soumission,
                      width=40, bg='#FEF9E7', font=('Arial', 10))
ent_list.grid(row=1, column=0)

frame_group_btn1 = tk.Frame(frame_soumission)
frame_group_btn1.grid(row=1, column=1, rowspan=4)

btn_1 = tk.Button(frame_group_btn1, text='>', font=('Arial', 11, 'bold'),
                  width=3, command=lambda: moveTo(ent_list, soum_list))
btn_1.grid(row=0, column=0, pady=5)

btn_2 = tk.Button(frame_group_btn1, text='>>', font=('Arial', 11, 'bold'),
                  width=3, command=lambda: move_all(ent_list, soum_list))
btn_2.grid(row=1, column=0, pady=5)

btn_3 = tk.Button(frame_group_btn1, text='<', font=('Arial', 11, 'bold'),
                  width=3, command=lambda: moveTo(soum_list, ent_list))
btn_3.grid(row=2, column=0, pady=5)

btn_4 = tk.Button(frame_group_btn1, text='<<', font=('Arial', 11, 'bold'),
                  width=3, command=lambda: move_all(soum_list, ent_list))
btn_4.grid(row=3, column=0, pady=5)


lbl_list_soum = tk.Label(frame_soumission, text='Liste des soumissionnaires')
lbl_list_soum.grid(row=0, column=2)
soum_list = tk.Listbox(frame_soumission, width=40,
                       bg='#FEF9E7', fg='#00F', font=('Arial', 10))
soum_list.grid(row=1, column=2)

frame_group_btn2 = tk.Frame(frame_soumission)
frame_group_btn2.grid(row=1, column=3)

btn_adj_1 = tk.Button(frame_group_btn2, text='>', state='disabled', font=('Arial', 11, 'bold'),
                      width=3, command=lambda: [move_adj(soum_list, adj_list), soum_to_adj])
btn_adj_1.grid(row=0, column=0, pady=5)

btn_adj_2 = tk.Button(frame_group_btn2, text='<', state='disabled', font=('Arial', 11, 'bold'),
                      width=3, command=lambda: [back_adj(adj_list, soum_list), adj_to_soum])
btn_adj_2.grid(row=1, column=0, pady=5)

lbl_adj = tk.Label(frame_soumission, text='Entreprise adjugée')
lbl_adj.grid(row=0, column=4)
adj_list = tk.Listbox(frame_soumission, width=40, bg='#FEF9E7',
                      fg='#1d5720', font=('Arial', 10))

for widget in frame_soumission.winfo_children():
    widget.grid_configure(padx=10, pady=0)

adj_list.grid(row=1, column=4, pady=5)

# ************************************************************************************************
# Remerciements
frame_remerc = tk.LabelFrame(root, text='Lettres de remerciement', fg='blue')
frame_remerc.grid(row=4, column=0, sticky='ew', padx=20, pady=5)

# Gabarit Remerciements
lbl_remerc = tk.Label(
    frame_remerc, text='Gabarit remerciement (.doc)', width=25)
lbl_remerc.grid(row=0, column=0)
btn_remerc = tk.Button(
    frame_remerc, text='Sélectionner...', command=select_remerc_file, width=15)
btn_remerc.grid(row=0, column=1)

# PV Ouverture Remerciements
lbl_pv_ouvert = tk.Label(
    frame_remerc, text="Procès verbal d'ouverture (.pdf)", width=25)
lbl_pv_ouvert.grid(row=0, column=2)
btn_pv_ouvert = tk.Button(
    frame_remerc, text='Sélectionner...', command=select_pv_ouverture_file, width=15)
btn_pv_ouvert.grid(row=0, column=3)

btn_gen_remerc = tk.Button(
    frame_remerc, text='Générer les lettres de remerciement', bg='#123456', fg='white', width=45, command=gener_remerc)
btn_gen_remerc.grid(row=0, column=4, sticky='w', padx=15)

for widget in frame_remerc.winfo_children():
    widget.grid_configure(padx=15, pady=10)

# Octroi
frame_octroi = tk.LabelFrame(root, text="Lettre d'octroi", fg='blue')
frame_octroi.grid(row=5, column=0, sticky='ew', padx=20, pady=5)

# Gabarit Octroi
lbl_octroi = tk.Label(frame_octroi, text="Gabarit d'octroi (.doc)", width=25)
lbl_octroi.grid(row=0, column=0)
btn_octroi = tk.Button(
    frame_octroi, text='Sélectionner...', command=select_octroi_file, width=15)
btn_octroi.grid(row=0, column=1)

# PV CA Octroi
lbl_pv_ca = tk.Label(frame_octroi, text='Procès verbal CA (.doc)', width=25)
lbl_pv_ca.grid(row=0, column=2)
btn_pv_ca = tk.Button(
    frame_octroi, text='Sélectionner...', command=select_pv_ca_file, width=15)
btn_pv_ca.grid(row=0, column=3)

btn_gen_octroi = tk.Button(
    frame_octroi, text="Générer la lettre d'octroi", bg='#123456', fg='white', width=45, command=gener_octroi)
btn_gen_octroi.grid(row=0, column=4, sticky="w", padx=15)

for widget in frame_octroi.winfo_children():
    widget.grid_configure(padx=15, pady=10)


# load_data()
# generate_publipostage()
ent_list.bind('<Double-Button>', dbl_moveTo)
soum_list.bind('<Double-Button>', dbl_moveBack)
soum_list.bind('<<ListboxSelect>>', soum_to_adj)
adj_list.bind('<<ListboxSelect>>', adj_to_soum)
cmb_secretaire.bind('<<ComboboxSelected>>', get_secret_name)

root.mainloop()